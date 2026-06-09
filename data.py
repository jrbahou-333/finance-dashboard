"""
data.py — loads real data from ECONOMIC EVALUATION.xlsx,
plus manual entries added via the app (stored alongside this file).
"""

import json
import os
import pandas as pd
import openpyxl

EXCEL_PATH = "ECONOMIC EVALUATION.xlsx"

# Personal data files — gitignored, never committed.
# sample_data/ versions are committed as runnable placeholders for anyone
# cloning without real data.
MANUAL_SNAPSHOTS_PATH = "manual_snapshots.csv"
ACCOUNTS_CONFIG_PATH  = "accounts_config.json"
ONE_OFF_EXPENSES_PATH = "one_off_expenses.csv"

SAMPLE_MANUAL_SNAPSHOTS_PATH = "sample_data/manual_snapshots.csv"
SAMPLE_ACCOUNTS_CONFIG_PATH  = "sample_data/accounts_config.json"
SAMPLE_ONE_OFF_EXPENSES_PATH = "sample_data/one_off_expenses.csv"

USING_REAL_EXCEL = os.path.exists(EXCEL_PATH)


def _resolve(real_path, sample_path):
    """Return real path if it exists, otherwise fall back to the sample."""
    return real_path if os.path.exists(real_path) else sample_path


def _load_workbook():
    if not USING_REAL_EXCEL:
        raise FileNotFoundError(
            f"'{EXCEL_PATH}' not found. Add your own spreadsheet or run in demo mode "
            "(the app uses sample_data/ files automatically when the Excel is absent)."
        )
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)


# ---------------------------------------------------------------------------
# NET WORTH SNAPSHOTS
# Sheet: Net Worth Input — "Reports - Breakdown" section
# Columns: Date, Account, Amount (rows after the header at row 9)
# ---------------------------------------------------------------------------
def _load_net_worth_snapshots():
    wb = _load_workbook()
    ws = wb["Net Worth Input"]
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "Date" and row[1] == "Account":
                header_found = True
            continue
        date, account, amount = row[0], row[1], row[2]
        if date is None and account is None:
            break
        if date is not None and account is not None and isinstance(amount, (int, float)):
            rows.append({"date": date, "account": account, "amount": amount})
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values("date")


# ---------------------------------------------------------------------------
# MONTHLY INCOME — from Calcs sheet (most recent income value)
# ---------------------------------------------------------------------------
def _load_monthly_income():
    wb = _load_workbook()
    ws = wb["Calcs"]
    # Find header row, then take latest income value
    header_found = False
    latest_income = None
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "Date" and row[4] == "Income":
                header_found = True
            continue
        if row[0] is None:
            break
        if isinstance(row[4], (int, float)):
            latest_income = row[4]
    return latest_income or 0


# ---------------------------------------------------------------------------
# MONTHLY EXPENSES — new expense section in Montly Expenses sheet
# ---------------------------------------------------------------------------
def _load_monthly_expenses():
    wb = _load_workbook()
    ws = wb["Montly Expenses"]
    rows = []
    in_new_section = False
    for row in ws.iter_rows(values_only=True):
        if row[1] == "Monthly Expenditure (new):":
            in_new_section = True
            continue
        if not in_new_section:
            continue
        # header row
        if row[1] == "Category":
            continue
        # total row or blank
        if row[1] is None or row[4] == "Total":
            if row[4] == "Total":
                break
            continue
        category, as_monthly = row[1], row[5]
        if category is not None and isinstance(as_monthly, (int, float)) and as_monthly > 0:
            rows.append({"category": category, "amount": as_monthly})
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# ONE-OFF EXPENSES
# Sheet: Other Expenditure — columns B:F
# ---------------------------------------------------------------------------
def _load_one_off_expenses():
    wb = _load_workbook()
    ws = wb["Other Expenditure"]
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        _, reason, category, date, amount, note = (list(row) + [None] * 6)[:6]
        if not header_found:
            if reason == "Reason":
                header_found = True
            continue
        if reason is None or not isinstance(amount, (int, float)):
            continue
        rows.append({"reason": reason, "category": category, "date": date, "amount": amount, "note": note})
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    return df


# ---------------------------------------------------------------------------
# ONE-OFF EXPENSES — editable store
# Seeded once from the Excel "Other Expenditure" sheet into one_off_expenses.csv,
# then managed entirely through the app (Excel is never modified).
# ---------------------------------------------------------------------------
_ONE_OFF_COLUMNS = ["reason", "category", "date", "amount", "note"]


def _seed_one_off_expenses():
    df = _load_one_off_expenses()
    df.to_csv(ONE_OFF_EXPENSES_PATH, index=False)
    return df


def load_one_off_expenses():
    path = _resolve(ONE_OFF_EXPENSES_PATH, SAMPLE_ONE_OFF_EXPENSES_PATH)
    if os.path.exists(path):
        df = pd.read_csv(path, parse_dates=["date"])
        df["note"] = df["note"].where(df["note"].notna(), None)
        return df[_ONE_OFF_COLUMNS]
    return _seed_one_off_expenses() if USING_REAL_EXCEL else pd.DataFrame(columns=_ONE_OFF_COLUMNS)


def _save_one_off_expenses(df):
    df = df[_ONE_OFF_COLUMNS].sort_values("date").reset_index(drop=True)
    df.to_csv(ONE_OFF_EXPENSES_PATH, index=False)
    return df


def add_one_off_expense(reason, category, date, amount, note=None):
    df = load_one_off_expenses()
    new_row = pd.DataFrame([{
        "reason": reason, "category": category,
        "date": pd.Timestamp(date), "amount": amount,
        "note": note or None,
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    _save_one_off_expenses(df)


def update_one_off_expense(row_id, reason, category, date, amount, note=None):
    df = load_one_off_expenses()
    if row_id not in df.index:
        return False
    df.loc[row_id] = {
        "reason": reason, "category": category,
        "date": pd.Timestamp(date), "amount": amount,
        "note": note or None,
    }
    _save_one_off_expenses(df)
    return True


def delete_one_off_expense(row_id):
    df = load_one_off_expenses()
    if row_id not in df.index:
        return False
    df = df.drop(index=row_id)
    _save_one_off_expenses(df)
    return True


# ---------------------------------------------------------------------------
# CASH FLOW — monthly income vs spend
# Sheet: Calcs
# ---------------------------------------------------------------------------
def _load_cash_flow():
    wb = _load_workbook()
    ws = wb["Calcs"]
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "Date" and row[1] == "Monthly Expenditure":
                header_found = True
            continue
        date, monthly_exp, one_off_exp, total_exp, income, remaining, cumulative = (list(row) + [None]*7)[:7]
        if date is None:
            break
        if not isinstance(monthly_exp, (int, float)):
            continue
        rows.append({
            "date": date,
            "monthly_expenses": monthly_exp or 0,
            "one_off_expenses": one_off_exp or 0,
            "income": income or 0,
            "remaining": remaining or 0,
            "cumulative": cumulative or 0,
        })
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    df["total_expenses"] = df["monthly_expenses"] + df["one_off_expenses"]
    return df


def get_cash_flow():
    """
    Cash flow recomputed live: recurring monthly expenses + the *editable*
    one-off expenses store, so adding/editing one-off expenses immediately
    flows through to the Cash Flow & Pinch Points charts. Historical income
    figures are kept from the Excel Calcs sheet (income varies month to
    month); months without a recorded income fall back to the latest known
    monthly income.
    """
    base = _load_cash_flow()
    oo = load_one_off_expenses()
    monthly_total = _load_monthly_expenses()["amount"].sum()

    base_periods = base["date"].dt.to_period("M")
    income_lookup = dict(zip(base_periods, base["income"]))
    latest_income = base["income"].iloc[-1] if len(base) else _load_monthly_income()

    oo_periods = oo["date"].dt.to_period("M") if len(oo) else pd.Series([], dtype="period[M]")
    one_off_by_month = oo.groupby(oo_periods)["amount"].sum() if len(oo) else pd.Series(dtype=float)

    months = sorted(set(base_periods) | set(oo_periods))
    rows = []
    for m in months:
        rows.append({
            "date": m.to_timestamp(),
            "monthly_expenses": monthly_total,
            "one_off_expenses": float(one_off_by_month.get(m, 0)),
            "income": income_lookup.get(m, latest_income),
        })
    df = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
    df["total_expenses"] = df["monthly_expenses"] + df["one_off_expenses"]
    df["remaining"] = df["income"] - df["total_expenses"]
    df["cumulative"] = df["remaining"].cumsum()
    return df


# ---------------------------------------------------------------------------
# PROJECTIONS — projected vs actual net worth
# Sheet: Projections
# ---------------------------------------------------------------------------
def _load_projections():
    wb = _load_workbook()
    ws = wb["Projections"]
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] == "Date" and row[3] == "Projected Worth":
                header_found = True
            continue
        date, _, _, projected, actual = (list(row) + [None]*5)[:5]
        if date is None:
            break
        if not isinstance(projected, (int, float)):
            continue
        rows.append({
            "date": date,
            "projected": projected,
            "actual": actual if isinstance(actual, (int, float)) else None,
        })
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    return df


# ---------------------------------------------------------------------------
# GOALS
# Sheet: GOALS
# ---------------------------------------------------------------------------
def _load_goals():
    wb = _load_workbook()
    ws = wb["GOALS"]
    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        _, goal, cost, months, monthly_saving, comment = (list(row) + [None]*6)[:6]
        if not header_found:
            if goal == "Goal":
                header_found = True
            continue
        if goal is None or not isinstance(cost, (int, float)):
            continue
        rows.append({
            "goal": goal,
            "target": cost,
            "months": months,
            "monthly_saving": monthly_saving or (cost / months if months else 0),
            "comment": comment,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# PENSION
# Sheet: PENSION
# ---------------------------------------------------------------------------
def _load_pension():
    wb = _load_workbook()
    ws = wb["PENSION"]
    data = {}
    rows = [(r[0], r[1]) for r in ws.iter_rows(max_col=2, values_only=True)]

    def find(label):
        for k, v in rows:
            if k == label:
                return v
        return None

    data["legal_general"] = {
        "name": "Legal & General — Abbott Retirement Saver",
        "invested": find("Invested"),
        "balance": find("Balance"),
        "date": str(find("Date"))[:10] if find("Date") else "—",
    }

    # Cushon section starts after "Cushon"
    cushon_idx = next((i for i, (k, _) in enumerate(rows) if k == "Cushon"), None)
    if cushon_idx is not None:
        cushon_rows = rows[cushon_idx:]
        def cfind(label):
            for k, v in cushon_rows:
                if k == label:
                    return v
            return None
        data["cushon"] = {
            "name": "Cushon — Aether Pension",
            "monthly_me": cfind("Me"),
            "monthly_employer": cfind("Aether"),
            "balance": cfind("Balance"),
            "date": str(cfind("Date"))[:10] if cfind("Date") else "—",
        }
    return data


# ---------------------------------------------------------------------------
# MANUAL SNAPSHOTS — monthly investment values added via the app
# Stored in manual_snapshots.csv: date, account, amount
# ---------------------------------------------------------------------------
def load_manual_snapshots():
    path = _resolve(MANUAL_SNAPSHOTS_PATH, SAMPLE_MANUAL_SNAPSHOTS_PATH)
    if os.path.exists(path):
        df = pd.read_csv(path, parse_dates=["date"])
        if not df.empty:
            return df
    return pd.DataFrame(columns=["date", "account", "amount"])


def add_manual_snapshot(date, account_amounts: dict):
    """Append a new monthly snapshot. account_amounts: {account_name: amount}"""
    df = load_manual_snapshots()
    new_rows = pd.DataFrame([
        {"date": pd.Timestamp(date), "account": acc, "amount": amt}
        for acc, amt in account_amounts.items()
    ])
    df = pd.concat([df, new_rows], ignore_index=True)
    df.to_csv(MANUAL_SNAPSHOTS_PATH, index=False)


def delete_manual_snapshot(date):
    """Remove a manually-added snapshot for a given date (only affects manual entries)."""
    df = load_manual_snapshots()
    df = df[pd.Timestamp(date) != pd.to_datetime(df["date"])]
    df.to_csv(MANUAL_SNAPSHOTS_PATH, index=False)


# ---------------------------------------------------------------------------
# ACCOUNT CONFIG — which investment accounts are tracked going forward
# Stored in accounts_config.json: {"active": [...], "removed": [...]}
# Seeded from whatever accounts already appear in the Excel net worth data.
# ---------------------------------------------------------------------------
def _seed_account_config():
    accounts = sorted(_load_net_worth_snapshots()["account"].dropna().unique().tolist())
    config = {"active": accounts, "removed": []}
    _save_account_config(config)
    return config


def _save_account_config(config):
    with open(ACCOUNTS_CONFIG_PATH, "w") as f:
        json.dump(config, f, indent=2)


def load_account_config():
    path = _resolve(ACCOUNTS_CONFIG_PATH, SAMPLE_ACCOUNTS_CONFIG_PATH)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return _seed_account_config()


def get_active_accounts():
    return load_account_config()["active"]


def add_account(name):
    config = load_account_config()
    name = name.strip()
    if not name:
        return False, "Account name can't be empty."
    if name in config["active"]:
        return False, f'"{name}" is already tracked.'
    if name in config["removed"]:
        config["removed"].remove(name)
    config["active"].append(name)
    _save_account_config(config)
    return True, f'Added "{name}".'


def remove_account(name):
    config = load_account_config()
    if name not in config["active"]:
        return False, f'"{name}" is not currently tracked.'
    config["active"].remove(name)
    config["removed"].append(name)
    _save_account_config(config)
    return True, f'Removed "{name}" from future tracking. Historical data is kept.'


# ---------------------------------------------------------------------------
# Public exports — these are what app.py uses
# ---------------------------------------------------------------------------
def get_net_worth_snapshots():
    """Excel snapshots merged with any manually-added monthly entries."""
    excel_df = _load_net_worth_snapshots()
    manual_df = load_manual_snapshots()
    combined = pd.concat([excel_df, manual_df], ignore_index=True)
    combined["date"] = pd.to_datetime(combined["date"])
    return combined.sort_values("date").reset_index(drop=True)


def get_one_off_expenses():
    """Editable one-off expenses (seeded from Excel, then app-managed)."""
    return load_one_off_expenses()


NET_WORTH_SNAPSHOTS = get_net_worth_snapshots()
MONTHLY_INCOME      = _load_monthly_income()
MONTHLY_EXPENSES    = _load_monthly_expenses()
ONE_OFF_EXPENSES    = get_one_off_expenses()
CASH_FLOW           = _load_cash_flow()
PROJECTIONS         = _load_projections()
GOALS               = _load_goals()
PENSION             = _load_pension()
